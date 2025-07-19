from getnav_func import get_nav
import pandas as pd
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
from tkinter import Tk


Tk().withdraw()
filename = askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])

if filename:
    print(f'\nUsing file {filename}')
    df = pd.read_excel(filename)

    # Process and collect NAV values
    nav_updates = {}
    for index, row in df.iterrows():
        url = row['URL']
        if pd.notna(url):
            print(f"Scraping NAV from: {url}")
            nav = get_nav(url)
            nav_updates[index] = nav
            # Update DataFrame for reference only
            df.at[index, 'Current NAV'] = nav
        else:
            print(f"Skipping row {index + 1}: No URL provided")



    # Now use openpyxl to update only the specific cells while preserving formulas
    wb = load_workbook(filename)
    ws = wb.active

    # Find the column index for 'Current NAV'
    header_row = next(ws.rows)
    nav_col_idx = None
    for idx, cell in enumerate(header_row, 1):
        if cell.value == 'Current NAV':
            nav_col_idx = idx
            break

    if nav_col_idx is not None:
        # Update only the NAV cells
        for idx, nav_value in nav_updates.items():
            # Excel rows are 1-based and have a header, so add 2 to the pandas index
            row_num = idx + 2
            ws.cell(row=row_num, column=nav_col_idx, value=nav_value)

        # Save the workbook
        wb.save("testnav.xlsx")
        print(f"Excel file updated successfully with the latest NAV values (formulas preserved)")
    else:
        print("Could not find 'Current NAV' column in the Excel file")
else:
    print("No Excel file name provided")






